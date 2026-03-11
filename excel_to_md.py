'''
Excel → Markdown (OpenPyXl 라이브러리)

기능 설명
- xlsx 확장자만 가능, xls 불가
- 엑셀 파일 내부 셀 테두리 기준으로 테이블 판별
- 테이블 내 빈 열/행 삭제
- #: 시트 제목, ##: 소제목, 단순 패턴 매칭(숫자 + '.' + 텍스트, 현재 제외)
- 시트 범위, 병합셀 등 옵션 설정 가능

옵션 설정
- MODE: 파일 범위 설정(절대 경로)
  ALL: 폴더 내 전체 파일, FILE: 단일 파일
- SHEET_MODE: 시트 설정
  ALL: 모든 시트, FIRST: 첫번째 시트만
- EXCLUDE_HIDDEN: 숨겨진 시트 설정
  True: 숨겨진 시트 제외, False: 숨겨진 시트 포함
- MERGED_MODE: 병합셀 처리
  SINGLE: 좌상단 한번만, EXPAND: 병합셀 전체에 값 반복
'''

from pathlib import Path
from openpyxl import load_workbook
from collections import deque
import re

# ===== 파일 처리 옵션 =====
MODE = "ALL"

# ALL 모드일 때 사용
ROOT = Path(r"D:\tchat_preprocess_repo\tchat_preprocessing\Tech_doc\Mat_Dept\basic_research\raw_data")

# FILE 모드일 때 사용
FILE = Path(r"D:\tchat_preprocess_repo\performance_test\test_file\C0000000000010455057_R0000000000010608265_160524_Dr.Holzbach Consulting 보고서.xlsx")
# "D:\tchat_preprocess_repo\tchat_preprocessing\Tech_doc\Mat_Dept\Compound\raw_data\C0000000000010455090_R0000000000010608299_2022_Y85(마모 향상 Alpin Studless) 개발.xlsx"
# "D:\tchat_preprocess_repo\tchat_preprocessing\Tech_doc\Mat_Dept\Compound\raw_data\C0000000000010462046_R0000000000010615527_중국 Tubeless 중단거리 내마모용 컴파운드 임시생산 결과 보고서 별첨.xlsx"
# "D:\tchat_preprocess_repo\performance_test\test_file\C0000000000010455057_R0000000000010608265_160524_Dr.Holzbach Consulting 보고서.xlsx"
# "D:\tchat_preprocess_repo\tchat_preprocessing\Tech_doc\Mat_Dept\basic_research\raw_data\C0000000000010461766_R0000000000010615246_[21년] The role of mercapto silane in silica reinforced SBR-BR compound_P55X 기반 new silane 평가 보고_최종.xlsx"

# 출력 폴더
OUTPUT_DIR = Path(r"D:\workspace\basic_research")

# ===== 시트 처리 옵션 =====
SHEET_MODE = "FIRST"
EXCLUDE_HIDDEN = True

# ===== 병합 셀 옵션 =====
MERGED_MODE = "EXPAND"


# 셀 테두리 여부 확인
def has_border(cell):
    b = cell.border
    return any([b.left.style, b.right.style, b.top.style, b.bottom.style])

# 셀 테두리 개수 확인
def count_border(cell):
    b = cell.border

    styles = [b.left.style, b.right.style, b.top.style, b.bottom.style]
    count = sum(1 for s in styles if s)

    return count >= 2

# 변환에 사용할 시트 설정
def select_sheets(wb):
    sheets = []
    if SHEET_MODE == "FIRST":
        for ws in wb.worksheets:
            if EXCLUDE_HIDDEN and ws.sheet_state == "visible":
                return [ws]
        return []
    elif SHEET_MODE == "ALL":
        for ws in wb.worksheets:
            if EXCLUDE_HIDDEN and ws.sheet_state != "visible":
                continue
            sheets.append(ws)
        return sheets
    else:
        raise ValueError("SHEET_MODE는 'ALL' 또는 'FIRST' 중 하나여야 합니다.")


# 소제목 태그 추가(제외)
def to_subheading(text):
    if re.match(r'^\s*\d+\..+', text):
        return f"## {text.strip()}"
    return text

# 테이블 좌표 반환
def find_table_blocks(ws):
    visited = set()
    blocks = []
    max_row = ws.max_row
    max_col = ws.max_column

    # 셀 순회
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            if (row, col) in visited:
                continue
            
            # 셀 테두리 여부 확인
            cell = ws.cell(row, col)
            if not has_border(cell):
                continue

            queue = deque([(row, col)])
            block = []

            while queue:
                r, c = queue.popleft()
                if (r, c) in visited:
                    continue

                current = ws.cell(r, c)
                if not has_border(current):
                    continue

                visited.add((r, c))
                block.append((r, c))

                # 인접 셀이 같은 표에 포함되는지 확인
                for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
                    nr, nc = r+dr, c+dc
                    if 1 <= nr <= max_row and 1 <= nc <= max_col:
                        if (nr, nc) not in visited:
                            queue.append((nr, nc))
            if block:
                blocks.append(block)
    return blocks

# 테이블 사이즈 확인
def block_bounds(block):
    rows = [r for r, _ in block]
    cols = [c for _, c in block]
    return min(rows), max(rows), min(cols), max(cols)

# 병합셀 설정
def build_merged_map(ws):
    merged_map = {}
    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col
        value = ws.cell(min_row, min_col).value

        if MERGED_MODE == "SINGLE":
            # 좌상단 셀만 값 기록, 나머지는 빈 값
            merged_map[(min_row, min_col)] = value
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if r == min_row and c == min_col:
                        continue
                    merged_map[(r, c)] = None
        elif MERGED_MODE == "EXPAND":
            # 병합셀 각각에 동일한 값 기록
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    merged_map[(r, c)] = value
        else:
            raise ValueError("MERGED_MODE 'SINGLE' 또는 'EXPAND' 중 하나여야 합니다.")
    return merged_map

# 테이블 → 마크다운 변환
def block_to_markdown(ws, block):
    min_r, max_r, min_c, max_c = block_bounds(block)

    md = []
    text = []
    prev_values = {}

    # 원본 기준 빈 열 삭제
    empty_cols = set()
    for c in range(min_c, max_c + 1):
        is_col_empty = True
        for r in range(min_r, max_r + 1):
            if ws.cell(r, c).value not in (None, ""):
                is_col_empty = False
                break
        if is_col_empty:
            empty_cols.add(c)

    merged_map = build_merged_map(ws)

    remained_cols = [c for c in range(min_c, max_c + 1) if c not in empty_cols]
    col_count = len(remained_cols)

    # 남는 열이 없으면 빈 문자열 반환
    if col_count == 0:
        return ""

    # 행 개수 계산
    row_count = 0
    for r in range(min_r, max_r + 1):
        is_row_empty = True
        for c in range(min_c, max_c + 1):
            if c in empty_cols:
                continue
            if ws.cell(r, c).value not in (None, ""):
                is_row_empty = False
                break
        if not is_row_empty:
            row_count += 1

    for r in range(min_r, max_r + 1):
        # 원본 기준 빈 행 삭제
        is_row_empty = True
        for c in range(min_c, max_c + 1):
            if c in empty_cols:
                continue
            if ws.cell(r, c).value not in (None, ""):
                is_row_empty = False
                break

        if is_row_empty:
            continue

        row_vals = []

        for c in range(min_c, max_c + 1):
            if c in empty_cols:
                continue

            val = merged_map.get((r, c), ws.cell(r, c).value)

            if val is not None:
                val = str(val).replace("\n", " ").replace("\r", " ")

            key = c

            if MERGED_MODE == "SINGLE":
                if key in prev_values and prev_values[key] == val:
                    row_vals.append("")
                else:
                    row_vals.append("" if val is None else str(val))
                    prev_values[key] = val
            else:
                row_vals.append("" if val is None else str(val))

        # 행이 1개 이하인 경우 텍스트로 반환
        if row_count <= 1:
            line_text = " ".join(v for v in row_vals if v)
            if line_text.strip():
                text.append(line_text.strip())
        else:
            md.append("| " + " | ".join(row_vals) + " |")

    # 행이 1개 이하인 경우 텍스트로 반환
    if row_count <= 1:
        return "\n".join(text)

    # 헤더 구분선 → 첫 행을 헤더로 가정
    if len(md) > 1:
        col_count_in_md = len(md[0].split("|")) - 2
        separator = "|" + "|".join(["---"] * col_count_in_md) + "|"
        md.insert(1, separator)

    return "\n".join(md)

# 엑셀 → 마크다운 변환
def convert_excel_to_md(input_path, output_path):
    if False:
        print(f"△ 이미 존재, 변환 생략: {output_path.name}")
        return

    try:
        wb = load_workbook(input_path, data_only=True)
        output_lines = []

        for ws in select_sheets(wb):
            sheet_name = ws.title
            output_lines.append(f"\n# {sheet_name}\n")

            # 테두리 기반 블록 탐색
            blocks = find_table_blocks(ws)
            # blocks = filter_blocks_by_first_cell_border(ws, blocks)

            # 외곽 프레임 처리: 경계 제거 후 내부 재분석
            # blocks = filter_outer_frame_blocks(ws, blocks)

            # 시작 행 기준으로 정렬
            blocks_sorted = sorted(blocks, key=lambda b: block_bounds(b)[0])

            # 시작 행 → (블록, 끝 행) 매핑
            block_map = {}
            for block in blocks_sorted:
                min_r, max_r, _, _ = block_bounds(block)
                if min_r not in block_map:
                    block_map[min_r] = []
                block_map[min_r].append((block, max_r))

            current_row = 1
            max_row = ws.max_row

            while current_row <= max_row:
                if current_row in block_map:
                    blocks_at_row = block_map[current_row]
                    
                    # 열 순서대로 정렬
                    blocks_at_row_sorted = sorted(blocks_at_row, key=lambda x: block_bounds(x[0])[2])
                    
                    for block, max_r in blocks_at_row_sorted:
                        min_r, _, _, _ = block_bounds(block)
                        first_row_cells = [(r, c) for (r, c) in block if r == min_r]
                        first_row_vals = [ws.cell(r, c).value for r, c in first_row_cells]
                        
                        # 첫 행에 값이 1개만 있으면 제목으로 처리
                        non_empty_vals = [v for v in first_row_vals if v not in (None, "")]
                        
                        if len(non_empty_vals) == 1:
                            # 첫 행을 제목으로 출력
                            title = str(non_empty_vals[0]).strip()
                            output_lines.append(f"\n{title}\n")
                            
                            # 첫 행 제거한 블록으로 테이블 생성
                            new_block = [(r, c) for (r, c) in block if r != min_r]
                            if new_block:
                                md_table = block_to_markdown(ws, new_block)
                                if md_table.strip():
                                    output_lines.append(md_table)
                                    output_lines.append("")
                        else:
                            # 일반 테이블
                            md_table = "\n" + block_to_markdown(ws, block)
                            if md_table.strip():
                                output_lines.append(md_table)
                                output_lines.append("")

                    max_end_row = max(end_r for _, end_r in blocks_at_row)
                    current_row = max_end_row + 1
                    continue

                # 테이블이 아닌 일반 텍스트 영역
                row_vals = [cell.value for cell in ws[current_row]]
                text = "\n".join(str(v) for v in row_vals if v)
                if text.strip():
                    # text = to_subheading(text)
                    output_lines.append(text)
                current_row += 1

        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(output_lines))

        print(f"✓ 변환 완료: {output_path}")
    except Exception as e:
        err_msg = str(e)
        print(f"✕ 변환 실패: {input_path.name}")
        print(f"내용: {err_msg[:200]}")


def collect_FILE():
    if MODE == "ALL":
        return list(ROOT.rglob("*.xlsx"))
    elif MODE == "FILE":
        if not FILE.exists():
            raise FileNotFoundError(f"파일이 존재하지 않습니다: {FILE}")
        return [FILE]
    else:
        raise ValueError("MODE는 'ALL' 또는 'FILE' 중 하나여야 합니다.")

def main():
    files = collect_FILE()
    for file_path in files:
        output_path = OUTPUT_DIR / (file_path.stem + ".md")
        convert_excel_to_md(file_path, output_path)

if __name__ == "__main__":
    main()
